import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

def append_data_to_excel(data_groups, file_path):
    # 检查文件是否存在
    file = Path(file_path)
    if file.is_file():
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active  # 默认获取活动sheet
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Data'

    # 找到第一个空白行
    row = sheet.max_row
    if sheet.cell(row=row, column=1).value is not None:
        row += 1  # 如果最后一行不为空，从下一行开始

    # 写入数据
    for group in data_groups:
        for col_index, value in enumerate(group):
            column_letter = get_column_letter(col_index + 1)
            sheet[f'{column_letter}{row}'] = value
        row += 1  # 每完成一组数据，行号增加

    # 保存工作簿
    workbook.save(file_path)

# 示例使用，可以根据需要修改或动态生成数据组
if __name__ == "__main__":
        data_groups = [
            [1, 3, 4, 6, 8, 10]  # 假设这次只添加这一组数据
        ]

        # 调用函数，将数据追加到现有的Excel文件
        append_data_to_excel(data_groups, "D:/document/MBAM/data/show_data/excel.xlsx")
        data_groups = [
            [2, 10]  # 假设这次只添加这一组数据
        ]
        append_data_to_excel(data_groups, "D:/document/MBAM/data/show_data/excel.xlsx")
