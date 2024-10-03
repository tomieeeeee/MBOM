#from pycallgraph import PyCallGraph
#from pycallgraph.output import GraphvizOutput
import sys
sys.path.append("D:/document/3v3")
from policy.MBAM import MBAM
from policy.MBAM_OM_MH import MBAM_OM_MH
from policy.MBAM_MH import MBAM_MH
from baselines.PPO import PPO, PPO_Buffer
from baselines.PPO_MH import PPO_MH, PPO_MH_Buffer
from baselines.PPO_OM_MH import PPO_OM_MH, PPO_OM_MH_Buffer
from env_wapper.simple_tag.simple_tag import Simple_Tag
import argparse
import time
from env_model.simple_tag.model_simple_tag import load_env_model as simple_tag_env_model
from env_model.simple_tag.model_simple_tag import ENV_Simple_Tag
if __name__ == '__main__':
  #graphviz = GraphvizOutput()
  #graphviz.output_file = 'render.png'
  #with PyCallGraph(output=graphviz):
    parser = argparse.ArgumentParser(description="test")
    parser.add_argument("--num_trj", type=int, default=3, help="Number of trajectories")
    parser.add_argument("--num_om_layers", type=int, default=5, help="Number of trajectories")
    parser.add_argument("--device", type=str, default="cpu", help="")
    parser.add_argument("--actor_rnn", type=bool, default=False, help="")
    parser.add_argument("--eps_per_epoch", type=int, default=1, help="")
    parser.add_argument("--save_per_epoch", type=int, default=5, help="")
    parser.add_argument("--true_prob", type=bool, default=True, help="True or False, edit Actor_RNN.py line 47-48")
    parser.add_argument("--prophetic_onehot", type=bool, default=False, help="True or False")
    parser.add_argument("--only_use_last_layer_IOP", type=bool, default=False, help="True or False")
    parser.add_argument("--random_best_response", type=bool, default=False, help="True or False")
    parser.add_argument("--record_more", type=bool, default=False, help="True or False")
    parser.add_argument("--rnn_mixer", type=bool, default=False, help="True or False")
    args = parser.parse_args()


    #file_dir = "/media/lenovo/144ED9814ED95C54/experiment_data/Simple_Tag/train/trueprob_simple_tag_ppo_vs_mbam_10oppo/0_1284741196/worker/0_2/model/"
    file_dir = "D:/document/MBAM/data/for_data/PPO_v10_1/set3/0_974160051/worker/0_0/model/"#D:\document\MBAM\data\for_data\PPO_v10_2\set3\\worker\0_0\model
    player1_file = file_dir + "PPO_MH_player1__player1_iter28900.ckp"
    #player2_file = file_dir + "PPO_MH_player2_player2_iter77200.ckp"
    player2_file = file_dir + "MBAM_player2_iter28900.ckp"
    #PPO_MH_player1__player1_iter120700
    player1_type = "ppo_mh" # "mbam_mh_om_mh"
    player2_type = "mbam_om_mh"

    
    env_model = simple_tag_env_model(args.device)
    player1_ctor = None
    player2_ctor = None
    agent1 = PPO_MH.load_model(filepath=player1_file, args=args, logger=None, device=args.device)
    #agent2 = PPO_MH.load_model(filepath=player2_file, args=args, logger=None, device=args.device)#, env_model=None
    agent2 = MBAM_OM_MH.load_model(filepath=player2_file, args=args, logger=None, device=args.device, env_model=env_model)
    env = Simple_Tag()
    agents = [agent1, agent2]
    score=[]
    score1=[]

    for i in range(100):
        dis_oppo = 0
        dis_me = 0
        temp=[]
        temp1=[]
        step = 0
        s = env.reset()
        while True:
            #env.render()
            #env.render("mode=rgb_array")
            #print(type(frame))
            action_info1 = []
            for i in range (10):
                action_info1.append(agent1.choose_action(state=s[i][0]))
            oppo_a = []
            for i in range (10):
                oppo_a.append([a.item() for a in action_info1[i][0]])
            
            #源代码a = agent2.choose_action(state=s[1], oppo_hidden_prob=action_info1[5])[0].item()
            start = time.clock()
            action_info2 = []
            
            for i in range (10):
                action_info2.append(agent2.choose_action(state=s[i][1]))
            
            a = []

            for i in range (10):
                a.append([a.item() for a in action_info2[i][0]])
            end = time.clock()
            print("解算时间",end-start)
            '''
            arrow_base = 0x2190 
            for i in range(6):
                if oppo_a[i] == 0:
                    oppo_action.append("-")
                else: 
                    offset = (oppo_a[i] - 1) % 4
                    oppo_action.append(chr(arrow_base+offset))
            '''
            
            
            actions = [oppo_a,a]
            for i in range(len(oppo_a)):
                for num in oppo_a[i]:
                    if num != 0 :
                        dis_oppo += 1
            for i in range(len(a)):
                for num in a[i]:
                    if num != 0 :
                        dis_me += 1
            #print("步数",step)
            s_, rew, done, _ = env.step(actions)
            temp.append(rew[0])
            temp1.append(rew[1])
            step+=1
    #print(rew)
    #if rew[0] >= 10:
    #    print("touch!!!!!!!")
            s = s_
            if done:
                print("jieshu__________第",i)
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                from pathlib import Path

                def append_data_to_excel(data_groups, file_path):
                    # 检查文件是否存在
                    file = Path(file_path)
                    if file.is_file():
                        workbook = openpyxl.load_workbook(file_path)
                        sheet = workbook.active  # 默认获取活动sheet
                    else:
                        workbook = Workbook()
                        sheet = workbook.active
                        sheet.title = 'Data'

                    # 找到第一个空白行
                    row = sheet.max_row
                    if sheet.cell(row=row, column=1).value is not None:
                        row += 1  # 如果最后一行不为空，从下一行开始

                    # 写入数据
                    for group in data_groups:
                        for col_index, value in enumerate(group):
                            column_letter = get_column_letter(col_index + 1)
                            sheet[f'{column_letter}{row}'] = value
                        row += 1  # 每完成一组数据，行号增加

                    # 保存工作簿
                    workbook.save(file_path)
                dis = [dis_oppo,dis_me]
                append_data_to_excel([dis], "D:/document/MBAM/data/show_data/dis.xlsx")
                #score_temp=sum(temp)/len(temp)
                #score_temp1=sum(temp1)/len(temp1)
                #score.append(score_temp)
                #score1.append(score_temp1)
                #
                # print(score)
                break
        '''
        import xlwt
        f = xlwt.Workbook('encoding = utf-8') #设置工作簿编码
        sheet1 = f.add_sheet('sheet1',cell_overwrite_ok=True) #创建sheet工作表
        #list1 = [1,3,4,6,8,10]#要写入的列表的值
        for i in range(len(score)):
             sheet1.write(i,0,score[i]) #写入数据参数对应 行, 列, 值
             sheet1.write(i,1,score1[i])
        f.save("D:/document/MBAM/data/show_data/text1.xls")#保存.xls到当前工作目录
        '''
